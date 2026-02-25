"""
app.py — TrafficAdmin Flask API v3
Endpoints:
  GET  /health
  POST /recognize
  POST /submit-violation   (face detect + violation classify + write DB)
  POST /classify-violation (text → category, no image needed)
  GET  /violations
  GET  /citizens
"""
from flask import Flask, request, jsonify
from flask_cors import CORS
import face_recognition, pickle, numpy as np
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import pandas as pd
import io, os
import requests

app = Flask(__name__)
CORS(app)

DB_FILE            = "traffic_db.xlsx"
ENCODINGS_FILE     = "encodings.pkl"
CLASSIFIER_FILE    = "violation_classifier.pkl"
EVIDENCE_DIR       = "evidence"
CONFIDENCE_THR     = 0.52

FINE_TABLE = {
    "Overspeeding":      1000,
    "Signal Jump":       500,
    "Drunk Driving":     5000,
    "No Helmet":         500,
    "Parking Violation": 300,
    "No License":        1000,
    "Wrong Lane":        300,
    "No Seatbelt":       500,
    "Triple Riding":     1000,
    "General":           500,
}

os.makedirs(EVIDENCE_DIR, exist_ok=True)

# ── Load face encodings ────────────────────────────────────────
print("\n[INFO] Loading face encodings...")
with open(ENCODINGS_FILE, "rb") as f:
    enc_data = pickle.load(f)
known_encodings = enc_data["encodings"]
known_names     = enc_data["names"]
print(f"[INFO] {len(known_encodings)} encodings for {len(set(known_names))} people.")

# ── Load violation classifier ──────────────────────────────────
print("[INFO] Loading violation classifier...")
with open(CLASSIFIER_FILE, "rb") as f:
    clf_data = pickle.load(f)
clf_pipeline = clf_data["pipeline"]
print(f"[INFO] Classifier ready. Classes: {clf_data['classes']}\n")


# ── Helpers ────────────────────────────────────────────────────
def classify_text(text):
    """Returns (category, confidence_pct)"""
    pred  = clf_pipeline.predict([text])[0]
    proba = float(clf_pipeline.predict_proba([text]).max()) * 100
    return pred, round(proba, 1)


def get_citizen(name):
    df = pd.read_excel(DB_FILE, sheet_name="Citizens", header=0)
    df.columns = df.columns.str.strip()
    m = df[df["Name"].astype(str).str.strip() == name.strip()]
    if m.empty: return None
    r = m.iloc[0]
    return {
        "name":             str(r.get("Name",               "")),
        "phone":            str(r.get("Phone Number",       "")),
        "aadhaar":          str(r.get("Aadhaar Number",     "")),
        "address":          str(r.get("Address",            "")),
        "license_no":       str(r.get("Driving License No", "")),
        "total_violations": int(r.get("Total Violations",   0)),
        "amount_due":       int(r.get("Amount Due (₹)",     0)),
        "status":           str(r.get("Status",             "")),
    }


def next_id(sheet, prefix, header=3):
    try:
        df = pd.read_excel(DB_FILE, sheet_name=sheet, header=header)
        df.columns = df.columns.str.strip()
        col = df.columns[0]
        ids  = df[col].dropna().astype(str).tolist()
        nums = [int(i.replace(prefix,"")) for i in ids if i.startswith(prefix)]
        return f"{prefix}{(max(nums)+1):03d}" if nums else f"{prefix}001"
    except Exception:
        return f"{prefix}001"


def do_face(img_bytes):
    pil = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    arr = np.array(pil, dtype=np.uint8)
    locs = face_recognition.face_locations(arr, model="hog")
    if not locs: return None, None, None, pil
    encs = face_recognition.face_encodings(arr, locs)
    if not encs: return None, None, None, pil
    dists = face_recognition.face_distance(known_encodings, encs[0])
    bi    = int(np.argmin(dists))
    bd    = float(dists[bi])
    conf  = round((1 - bd) * 100, 2)
    name  = known_names[bi] if bd <= CONFIDENCE_THR else None
    return name, conf, bd, pil


def cell_w(cell, val, bg="FFFFFF", center=False, bold=False, color="1A1A1A"):
    cell.value     = val
    cell.font      = Font(name="Arial", size=9, bold=bold, color=color)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border    = Border(
        left=Side(style="thin",  color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin",color="CCCCCC"))


# ══════════════════════════════════════════════════════════════
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok",
                    "people_loaded": len(set(known_names)),
                    "total_encodings": len(known_encodings),
                    "classifier_classes": clf_data["classes"]})


# ══════════════════════════════════════════════════════════════
@app.route("/classify-violation", methods=["POST"])
def classify_violation():
    """
    Classify free-text violation description into a category.
    Body JSON: { "text": "driver was drunk and swerving" }
    Returns:   { category, confidence, fine_amount }
    """
    body = request.get_json(silent=True)
    if not body or "text" not in body:
        return jsonify({"success": False, "error": "JSON with 'text' required"}), 400
    text = body["text"].strip()
    if not text:
        return jsonify({"success": False, "error": "Empty text"}), 400

    category, confidence = classify_text(text)
    fine = FINE_TABLE.get(category, 500)
    return jsonify({
        "success":    True,
        "input_text": text,
        "category":   category,
        "confidence": confidence,
        "fine_amount": fine,
    })


# ══════════════════════════════════════════════════════════════
@app.route("/recognize", methods=["POST"])
def recognize():
    if "image" not in request.files:
        return jsonify({"success": False, "error": "No image"}), 400
    try:
        name, conf, dist, _ = do_face(request.files["image"].read())
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

    if name is None and conf is None:
        return jsonify({"success": True, "matched": False, "message": "No face detected."})
    if name is None:
        return jsonify({"success": True, "matched": False,
                        "confidence": conf, "message": "Not in database."})
    return jsonify({"success": True, "matched": True, "name": name,
                    "confidence": conf, "citizen": get_citizen(name),
                    "detected_at": datetime.now().strftime("%d-%b-%Y %H:%M:%S")})


# ══════════════════════════════════════════════════════════════
@app.route("/submit-violation", methods=["POST"])
def submit_violation():
    """
    Accepts image + violation_text (free text) + place + city + date_time.
    Classifies violation text → category automatically.
    Detects face → looks up citizen.
    Writes Violation Record + Detection Log + updates Citizen totals.
    """
    if "image" not in request.files:
        return jsonify({"success": False, "error": "No image"}), 400

    place          = request.form.get("place",          "Unknown Location")
    city           = request.form.get("city",           "Unknown City")
    violation_text = request.form.get("violation_text", "General violation")
    date_time_str  = request.form.get("date_time",      datetime.now().strftime("%d-%b-%Y %H:%M"))

    # ── Classify violation text ────────────────────────────────
    violation_type, clf_confidence = classify_text(violation_text)
    fine_amount = FINE_TABLE.get(violation_type, 500)

    # ── Face recognition ───────────────────────────────────────
    img_bytes = request.files["image"].read()
    try:
        name, face_conf, dist, pil_img = do_face(img_bytes)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

    if name is None and face_conf is None:
        return jsonify({"success": False, "matched": False, "error": "No face detected."})
    if name is None:
        return jsonify({"success": False, "matched": False,
                        "face_confidence": face_conf,
                        "error": "Person not recognised.",
                        "violation_classified_as": violation_type})

    citizen = get_citizen(name)
    if not citizen:
        return jsonify({"success": False, "error": f"No record for {name}"}), 404

    # ── Save evidence ──────────────────────────────────────────
    vio_id = next_id("Violation Records", "VIO")
    log_id = next_id("Detection Log",     "LOG")
    ev_path = os.path.join(EVIDENCE_DIR, f"{vio_id}_{name.replace(' ','_')}.png")
    pil_img.save(ev_path)

    now_str = datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    wb      = load_workbook(DB_FILE)

    # Violation Records row
    ws_v = wb["Violation Records"]
    nr   = ws_v.max_row + 1
    for ci, val in enumerate([vio_id, name, city, place, violation_type,
                               date_time_str, fine_amount, "Unpaid", ev_path], start=1):
        cell_w(ws_v.cell(nr, ci), val, bg="FDECEA",
               center=(ci in [1,3,5,6,7,8]),
               bold=(ci==8), color="C0392B" if ci==8 else "1A1A1A")

    # Detection Log row
    ws_l = wb["Detection Log"]
    lr   = ws_l.max_row + 1
    for ci, val in enumerate([log_id, vio_id, name, citizen["aadhaar"],
                               citizen["phone"], city, place,
                               violation_type, f"{face_conf:.1f}%", now_str], start=1):
        cell_w(ws_l.cell(lr, ci), val, bg="EBF5FB", center=True)

    # Update Citizen totals
    ws_c = wb["Citizens"]
    for row in ws_c.iter_rows(min_row=5, max_row=ws_c.max_row):
        if row[0].value and str(row[0].value).strip() == name.strip():
            row[5].value = int(row[5].value or 0) + 1
            row[6].value = int(row[6].value or 0) + fine_amount
            total  = int(row[5].value)
            status = "Blacklisted" if total >= 5 else ("Pending" if total >= 2 else "Clear")
            sc     = "C0392B" if status=="Blacklisted" else ("B7950B" if status=="Pending" else "1E8449")
            row[7].value = status
            row[7].font  = Font(name="Arial", size=9, bold=True, color=sc)
            break

    wb.save(DB_FILE)

    try:
        r = requests.post(
            "https://purple-mud-c6ef.pratyush-gupta.workers.dev/",
            json={
                "id": vio_id,
                "name": name,
                "aadhaar": citizen["aadhaar"],
                "phone": citizen["phone"],
                "license_no": citizen["license_no"],
                "violation_type": violation_type,
                "place": place,
                "city": city,
                "date_time": date_time_str,
                "fine_amount": fine_amount,
            },
            timeout=5
        )

        print("[D1] Worker response:", r.text)

    except Exception as e:
        print("[D1 ERROR]", e)

    return jsonify({
        "success":            True,
        "matched":            True,
        "violation_id":       vio_id,
        "log_id":             log_id,
        "name":               name,
        "face_confidence":    face_conf,
        "violation_text":     violation_text,
        "violation_type":     violation_type,
        "clf_confidence":     clf_confidence,
        "fine_amount":        fine_amount,
        "city":               city,
        "place":              place,
        "date_time":          date_time_str,
        "evidence":           ev_path,
        "citizen":            citizen,
    })


# ══════════════════════════════════════════════════════════════
@app.route("/violations", methods=["GET"])
def get_violations():
    try:
        df = pd.read_excel(DB_FILE, sheet_name="Violation Records", header=3)
        df.columns = df.columns.str.strip()
        return jsonify({"success": True,
                        "violations": df.dropna(how="all").fillna("").to_dict("records")})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/citizens", methods=["GET"])
def get_citizens():
    try:
        df = pd.read_excel(DB_FILE, sheet_name="Citizens", header=0)
        df.columns = df.columns.str.strip()
        print(df.columns.tolist())
        return jsonify({"success": True,
                        "citizens": df.dropna(how="all").fillna("").to_dict("records")})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == "__main__":
    print("┌─────────────────────────────────────────┐")
    print("│   TrafficAdmin AI — Flask API v3        │")
    print("│   http://localhost:5000                 │")
    print("├─────────────────────────────────────────┤")
    print("│  GET  /health                           │")
    print("│  POST /classify-violation               │")
    print("│  POST /recognize                        │")
    print("│  POST /submit-violation                 │")
    print("│  GET  /violations                       │")
    print("│  GET  /citizens                         │")
    print("└─────────────────────────────────────────┘\n")
    app.run(host="0.0.0.0", port=5000, debug=True)